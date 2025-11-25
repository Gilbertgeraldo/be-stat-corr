# app/routers/ml.py - Ultra-lightweight untuk Vercel
from fastapi import APIRouter, HTTPException, File, UploadFile
import json
import csv
import io
from math import sqrt

router = APIRouter(tags=["Machine Learning"])

MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB

@router.post("/analyze/correlation-upload")
async def analyze_correlation_upload(file: UploadFile = File(...)):
    """
    Analisis korelasi TANPA pandas - Ultra lightweight untuk Vercel
    """
    try:
        # Validasi tipe file
        allowed_extensions = ['.xlsx', '.xls', '.csv']
        file_ext = None
        for ext in allowed_extensions:
            if file.filename.lower().endswith(ext):
                file_ext = ext
                break
        
        if not file_ext:
            raise HTTPException(
                status_code=400, 
                detail=f"Format file tidak didukung. Gunakan: CSV atau Excel"
            )
        
        # Baca file
        contents = await file.read()
        
        if len(contents) > MAX_FILE_SIZE:
            raise HTTPException(
                status_code=413,
                detail=f"File terlalu besar. Max: 5MB"
            )
        
        if not contents:
            raise HTTPException(status_code=400, detail="File kosong")
        
        # Parse file
        try:
            if file_ext == '.csv':
                data = parse_csv(contents)
            else:
                data = parse_excel(contents)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Gagal membaca file: {str(e)}")
        
        if not data:
            raise HTTPException(status_code=400, detail="File tidak memiliki data")
        
        # Ambil hanya kolom numerik
        numeric_data = extract_numeric_columns(data)
        
        if not numeric_data:
            raise HTTPException(
                status_code=400, 
                detail=f"Tidak ada kolom numerik dalam file"
            )
        
        # Limit untuk Vercel
        if len(numeric_data) > 1000:
            numeric_data = numeric_data[:1000]
        
        # Hitung korelasi
        correlation_matrix = calculate_correlation_matrix(numeric_data)
        p_values = calculate_p_values_simple(numeric_data)
        
        # Cari korelasi kuat
        strong_correlations = find_strong_correlations(
            numeric_data,
            correlation_matrix,
            p_values,
            threshold=0.7
        )
        
        # Summary stats
        summary_stats = calculate_summary_stats(numeric_data)
        
        return {
            "status": "success",
            "file_name": file.filename,
            "total_rows": len(numeric_data),
            "total_columns": len(numeric_data[0]) if numeric_data else 0,
            "numeric_columns": list(numeric_data[0].keys()) if numeric_data else [],
            "non_numeric_columns": [],
            "pearson_correlation": correlation_matrix,
            "p_values": p_values,
            "strong_correlations": strong_correlations,
            "data_types": {},
            "summary_stats": summary_stats
        }

    except HTTPException as e:
        raise e
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


def parse_csv(contents: bytes) -> list:
    """Parse CSV file"""
    try:
        text = contents.decode('utf-8')
        reader = csv.DictReader(io.StringIO(text))
        return list(reader)
    except:
        text = contents.decode('latin-1')
        reader = csv.DictReader(io.StringIO(text))
        return list(reader)


def parse_excel(contents: bytes) -> list:
    """Parse Excel file - minimal"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        # Ambil header
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # Ambil data
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {headers[i]: row[i] for i in range(len(headers))}
            data.append(row_dict)
        
        return data
    except:
        raise Exception("Gagal membaca Excel file")


def extract_numeric_columns(data: list) -> list:
    """Extract hanya kolom numerik"""
    if not data:
        return []
    
    numeric_data = []
    numeric_keys = set()
    
    # Identifikasi kolom numerik
    for row in data:
        for key, value in row.items():
            if value is None:
                continue
            try:
                float(value)
                numeric_keys.add(key)
            except (ValueError, TypeError):
                pass
    
    # Buat data dengan hanya kolom numerik
    for row in data:
        numeric_row = {}
        for key in numeric_keys:
            try:
                numeric_row[key] = float(row.get(key, 0))
            except:
                numeric_row[key] = 0.0
        if numeric_row:
            numeric_data.append(numeric_row)
    
    return numeric_data


def calculate_correlation_matrix(data: list) -> dict:
    """Hitung correlation matrix tanpa pandas"""
    if not data:
        return {}
    
    keys = list(data[0].keys())
    correlations = {}
    
    for key1 in keys:
        correlations[key1] = {}
        for key2 in keys:
            if key1 == key2:
                correlations[key1][key2] = 1.0
            else:
                corr = calculate_pearson(data, key1, key2)
                correlations[key1][key2] = round(corr, 4)
    
    return correlations


def calculate_pearson(data: list, key1: str, key2: str) -> float:
    """Hitung Pearson correlation"""
    n = len(data)
    if n < 2:
        return 0.0
    
    x = [row[key1] for row in data]
    y = [row[key2] for row in data]
    
    # Calculate means
    mean_x = sum(x) / n
    mean_y = sum(y) / n
    
    # Calculate covariance and standard deviations
    cov = sum((x[i] - mean_x) * (y[i] - mean_y) for i in range(n)) / (n - 1)
    std_x = sqrt(sum((xi - mean_x) ** 2 for xi in x) / (n - 1))
    std_y = sqrt(sum((yi - mean_y) ** 2 for yi in y) / (n - 1))
    
    if std_x == 0 or std_y == 0:
        return 0.0
    
    return cov / (std_x * std_y)


def calculate_p_values_simple(data: list) -> dict:
    """Estimasi p-value simple (tanpa scipy)"""
    if not data:
        return {}
    
    keys = list(data[0].keys())
    p_values = {}
    
    for key1 in keys:
        p_values[key1] = {}
        for key2 in keys:
            if key1 == key2:
                p_values[key1][key2] = 0.0
            else:
                # Simple estimation - correlation strength
                p_val = estimate_p_value_simple(data, key1, key2)
                p_values[key1][key2] = round(p_val, 6)
    
    return p_values


def estimate_p_value_simple(data: list, key1: str, key2: str) -> float:
    """Estimasi p-value simple berdasarkan correlation strength"""
    n = len(data)
    r = calculate_pearson(data, key1, key2)
    
    # Simple t-test estimation
    if n < 3:
        return 1.0
    
    try:
        t_stat = r * sqrt(n - 2) / sqrt(1 - r * r) if r * r < 1 else 0
        # Rough approximation
        p_value = 1.0 / (1.0 + abs(t_stat))
        return min(p_value, 1.0)
    except:
        return 0.5


def find_strong_correlations(data: list, corr_matrix: dict, p_values: dict, threshold: float = 0.7) -> list:
    """Cari korelasi kuat"""
    strong = []
    keys = list(corr_matrix.keys())
    
    for i, key1 in enumerate(keys):
        for key2 in keys[i+1:]:
            corr_val = corr_matrix[key1][key2]
            p_val = p_values[key1][key2]
            
            if abs(corr_val) > threshold and p_val < 0.05:
                strong.append({
                    "column_1": key1,
                    "column_2": key2,
                    "correlation": corr_val,
                    "p_value": p_val,
                    "strength": "Sangat Kuat" if abs(corr_val) > 0.9 else "Kuat",
                    "direction": "Positif" if corr_val > 0 else "Negatif"
                })
    
    return sorted(strong, key=lambda x: abs(x["correlation"]), reverse=True)


def calculate_summary_stats(data: list) -> dict:
    """Hitung summary statistics"""
    if not data:
        return {}
    
    stats = {}
    keys = list(data[0].keys())
    
    for key in keys:
        values = [row[key] for row in data if row[key] is not None]
        
        if not values:
            continue
        
        n = len(values)
        mean = sum(values) / n
        var = sum((x - mean) ** 2 for x in values) / (n - 1) if n > 1 else 0
        std = sqrt(var) if var >= 0 else 0
        
        stats[key] = {
            "mean": round(mean, 2),
            "std": round(std, 2),
            "min": round(min(values), 2),
            "max": round(max(values), 2),
            "25%": round(sorted(values)[int(n * 0.25)], 2) if n > 0 else 0,
            "50%": round(sorted(values)[int(n * 0.50)], 2) if n > 0 else 0,
            "75%": round(sorted(values)[int(n * 0.75)], 2) if n > 0 else 0,
            "count": n
        }
    
    return stats


@router.get("/analyze/correlation-info")
def correlation_info():
    """Informasi tentang analisis korelasi"""
    return {
        "info": "Analisis Korelasi Pearson - Lightweight",
        "interpretasi": {
            "1.0": "Korelasi sempurna positif",
            "0.7-0.99": "Korelasi sangat kuat positif",
            "0.5-0.69": "Korelasi kuat positif",
            "0.3-0.49": "Korelasi sedang positif",
            "0.1-0.29": "Korelasi lemah positif",
            "0.0-0.09": "Tidak ada korelasi",
            "-0.09-0.0": "Tidak ada korelasi",
            "-0.29--0.1": "Korelasi lemah negatif",
            "-0.49--0.3": "Korelasi sedang negatif",
            "-0.69--0.5": "Korelasi kuat negatif",
            "-0.99--0.7": "Korelasi sangat kuat negatif",
            "-1.0": "Korelasi sempurna negatif"
        }
    }